"""
Genera documento de comparacion Modelo vs Analista (estilo Toto).

Lee comparaciones/input_comparaciones.yaml, corre el modelo para cada combinacion,
extrae el parrafo_resumen, y genera:
  - comparaciones/comparacion_modelo_vs_analista.xlsx
  - comparaciones/comparacion_visual.html

Uso:
  python scripts/generar_comparacion.py              # Solo genera Excel/HTML con datos existentes
  python scripts/generar_comparacion.py --correr      # Corre el modelo para cada combinacion primero
"""

import sys
import os
import re
import json
import time
import subprocess
from pathlib import Path

script_dir = Path(__file__).parent.absolute()
project_root = script_dir.parent
sys.path.insert(0, str(project_root / "src"))

import yaml
import pandas as pd


def strip_html(text: str) -> str:
    """Remove HTML tags from text."""
    return re.sub(r"<[^>]+>", "", text or "").strip()


def load_razonamiento(site: str, fecha_final: str) -> dict | None:
    """Load razonamiento from checkpoint data."""
    data_dir = project_root / "data"
    cp1_path = data_dir / f"checkpoint1_consolidado_{site}_{fecha_final}.json"
    if not cp1_path.exists():
        return None

    with open(cp1_path, "r", encoding="utf-8") as f:
        cp1 = json.load(f)

    cp5_path = data_dir / f"checkpoint5_causas_raiz_{site}_{fecha_final}.json"
    cp5 = None
    if cp5_path.exists():
        with open(cp5_path, "r", encoding="utf-8") as f:
            cp5 = json.load(f)

    # Load config
    with open(project_root / "config" / "config.yaml", "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)

    # Load parquet
    parquet_path = data_dir / f"datos_nps_enriquecido_{site}_{fecha_final}.parquet"
    if not parquet_path.exists():
        parquet_path = data_dir / f"datos_nps_{site}_{fecha_final}.parquet"
    if not parquet_path.exists():
        return None

    df = pd.read_parquet(parquet_path)
    df = df[df["SITE"] == site].copy()

    update_tipo = config.get("update", {}).get("tipo", "all")
    if update_tipo != "all":
        from nps_model.analysis.updates import filtrar_por_update
        df = filtrar_por_update(df, update_tipo)

    from nps_model.analysis.razonamiento import ejecutar_razonamiento
    quarter_actual = config.get("quarter_actual", "")
    quarter_anterior = config.get("quarter_anterior", "")

    r = ejecutar_razonamiento(
        cp1, None, None, cp5, df, site, fecha_final, config,
        quarter_actual, quarter_anterior,
    )
    return r


def update_config(site: str, q_ant: str, q_act: str, update_tipo: str, fecha_corte: str = None):
    """Update config.yaml with the specified parameters."""
    config_path = project_root / "config" / "config.yaml"
    with open(config_path, "r", encoding="utf-8") as f:
        content = f.read()
        config = yaml.safe_load(content)

    # Update values using string replacement for safety
    with open(config_path, "r", encoding="utf-8") as f:
        text = f.read()

    # Sites
    old_sites = config.get("sites", ["MLB"])
    old_site_line = f"  - {old_sites[0]}" if old_sites else "  - MLB"
    text = text.replace(old_site_line, f"  - {site}", 1)

    # Quarters
    old_q_act = config.get("quarter_actual", "")
    old_q_ant = config.get("quarter_anterior", "")
    text = text.replace(f'quarter_actual: "{old_q_act}"', f'quarter_actual: "{q_act}"')
    text = text.replace(f'quarter_anterior: "{old_q_ant}"', f'quarter_anterior: "{q_ant}"')

    # Update tipo
    old_tipo = config.get("update", {}).get("tipo", "all")
    text = text.replace(f'tipo: "{old_tipo}"', f'tipo: "{update_tipo}"')

    # Fecha de corte
    import re
    if fecha_corte:
        text = re.sub(r'fecha_corte:.*', f'fecha_corte: "{fecha_corte}"', text)
    else:
        text = re.sub(r'fecha_corte:.*', 'fecha_corte: null', text)

    with open(config_path, "w", encoding="utf-8") as f:
        f.write(text)


def run_model():
    """Run the full model."""
    result = subprocess.run(
        [sys.executable, str(project_root / "ejecutar_modelo_completo.py")],
        capture_output=True, text=True, encoding="utf-8",
        cwd=str(project_root), timeout=1200,
        env={**os.environ, "PYTHONIOENCODING": "utf-8"},
    )
    return result.returncode == 0, result.stdout + result.stderr


def generate_excel(results: list, output_path: Path):
    """Generate Excel file with comparison data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Sellers"

    # Headers
    headers = ["Update", "Producto", "Site", "Variacion NPS", "Explicacion Analista",
               "Explicacion Modelo", "Convalida?", "Comment"]
    header_fill = PatternFill(start_color="FFE600", end_color="FFE600", fill_type="solid")
    header_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    convalida_fills = {
        "Si": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Direccional": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "No": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    for i, r in enumerate(results, 2):
        quarter = f"{r['quarter_anterior']} vs {r['quarter_actual']}"
        ws.cell(row=i, column=1, value=quarter).border = thin_border
        ws.cell(row=i, column=2, value=r["update_tipo"]).border = thin_border
        ws.cell(row=i, column=3, value=r["site"]).border = thin_border

        var_cell = ws.cell(row=i, column=4, value=r.get("variacion_nps"))
        var_cell.border = thin_border
        var_cell.number_format = "+0.0;-0.0;0.0"
        if r.get("variacion_nps") is not None:
            var_cell.font = Font(
                color="388E3C" if r["variacion_nps"] >= 0 else "D32F2F",
                bold=True,
            )

        ws.cell(row=i, column=5, value=r.get("explicacion_analista", "")).border = thin_border
        ws.cell(row=i, column=5).alignment = Alignment(wrap_text=True)

        ws.cell(row=i, column=6, value=r.get("explicacion_modelo", "")).border = thin_border
        ws.cell(row=i, column=6).alignment = Alignment(wrap_text=True)

        convalida = r.get("convalida", "")
        c_cell = ws.cell(row=i, column=7, value=convalida or "")
        c_cell.border = thin_border
        c_cell.alignment = Alignment(horizontal="center")
        if convalida in convalida_fills:
            c_cell.fill = convalida_fills[convalida]

        ws.cell(row=i, column=8, value=r.get("comment", "")).border = thin_border
        ws.cell(row=i, column=8).alignment = Alignment(wrap_text=True)

    # Column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 60
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 40

    ws.auto_filter.ref = f"A1:H{len(results)+1}"
    wb.save(output_path)
    print(f"   Excel guardado: {output_path}")


def generate_html(results: list, output_path: Path):
    """Generate HTML visual comparison."""
    import html as html_mod

    def esc(t):
        return html_mod.escape(str(t)) if t else ""

    # Summary stats
    total = len(results)
    si = sum(1 for r in results if r.get("convalida") == "Si")
    dir_ = sum(1 for r in results if r.get("convalida") == "Direccional")
    no = sum(1 for r in results if r.get("convalida") == "No")
    pending = total - si - dir_ - no

    html = f"""<!DOCTYPE html>
<html><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Comparacion Modelo vs Analista — NPS Relacional Sellers</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif; background:#f5f5f5; padding:24px; }}
  .container {{ max-width:1400px; margin:0 auto; }}
  h1 {{ font-size:22px; margin-bottom:8px; }}
  .summary {{ display:flex; gap:12px; margin:16px 0 24px; }}
  .summary-card {{ padding:12px 20px; border-radius:8px; font-size:14px; font-weight:600; }}
  .card-si {{ background:#C6EFCE; color:#388E3C; }}
  .card-dir {{ background:#FFEB9C; color:#E65100; }}
  .card-no {{ background:#FFC7CE; color:#C62828; }}
  .card-pending {{ background:#E3F2FD; color:#1565C0; }}
  table {{ width:100%; border-collapse:collapse; background:white; border-radius:10px; overflow:hidden; box-shadow:0 1px 4px rgba(0,0,0,.1); }}
  th {{ background:#FFE600; padding:10px 12px; font-size:12px; text-align:left; border-bottom:2px solid #ddd; }}
  td {{ padding:10px 12px; font-size:12px; border-bottom:1px solid #eee; vertical-align:top; }}
  tr:hover {{ background:#fafafa; }}
  .var-up {{ color:#388E3C; font-weight:700; }}
  .var-down {{ color:#D32F2F; font-weight:700; }}
  .badge-si {{ background:#C6EFCE; color:#388E3C; padding:2px 8px; border-radius:4px; font-weight:600; font-size:11px; }}
  .badge-dir {{ background:#FFEB9C; color:#E65100; padding:2px 8px; border-radius:4px; font-weight:600; font-size:11px; }}
  .badge-no {{ background:#FFC7CE; color:#C62828; padding:2px 8px; border-radius:4px; font-weight:600; font-size:11px; }}
  .badge-pending {{ background:#E3F2FD; color:#1565C0; padding:2px 8px; border-radius:4px; font-weight:600; font-size:11px; }}
  .expand {{ cursor:pointer; color:#1565C0; font-size:11px; }}
  .detail {{ display:none; margin-top:8px; padding:8px; background:#f9f9f9; border-radius:6px; font-size:11px; line-height:1.5; }}
  .detail.open {{ display:block; }}
</style>
</head><body>
<div class="container">
  <h1>Comparacion Modelo vs Analista — NPS Relacional Sellers</h1>
  <p style="color:#888;font-size:13px;margin-bottom:16px;">Generado: {time.strftime("%Y-%m-%d %H:%M")}</p>

  <div class="summary">
    <div class="summary-card card-si">Si: {si}/{total}</div>
    <div class="summary-card card-dir">Direccional: {dir_}/{total}</div>
    <div class="summary-card card-no">No: {no}/{total}</div>
    <div class="summary-card card-pending">Pendiente: {pending}/{total}</div>
  </div>

  <table>
    <thead>
      <tr>
        <th>Update</th><th>Producto</th><th>Site</th><th>Var NPS</th>
        <th style="width:30%;">Explicacion Analista</th>
        <th style="width:30%;">Explicacion Modelo</th>
        <th>Convalida?</th>
      </tr>
    </thead>
    <tbody>
"""

    for i, r in enumerate(results):
        quarter = f"{r['quarter_anterior']} vs {r['quarter_actual']}"
        var = r.get("variacion_nps")
        var_str = f"{var:+.1f}pp" if var is not None else "--"
        var_cls = "var-up" if var and var >= 0 else "var-down"

        convalida = r.get("convalida", "")
        badge_cls = {"Si": "badge-si", "Direccional": "badge-dir", "No": "badge-no"}.get(convalida, "badge-pending")
        badge_text = convalida or "Pendiente"

        analista = esc(r.get("explicacion_analista", "") or "")
        modelo = esc(r.get("explicacion_modelo", "") or "")

        # Truncate for display, full in expandable
        analista_short = analista[:200] + "..." if len(analista) > 200 else analista
        modelo_short = modelo[:200] + "..." if len(modelo) > 200 else modelo

        html += f"""      <tr>
        <td>{esc(quarter)}</td>
        <td><b>{esc(r['update_tipo'])}</b></td>
        <td>{esc(r['site'])}</td>
        <td class="{var_cls}">{var_str}</td>
        <td>{analista_short}{'<span class="expand" onclick="this.nextElementSibling.classList.toggle(\\x27open\\x27)"> [ver mas]</span><div class="detail">' + analista + '</div>' if len(analista) > 200 else ''}</td>
        <td>{modelo_short}{'<span class="expand" onclick="this.nextElementSibling.classList.toggle(\\x27open\\x27)"> [ver mas]</span><div class="detail">' + modelo + '</div>' if len(modelo) > 200 else ''}</td>
        <td><span class="{badge_cls}">{badge_text}</span></td>
      </tr>
"""

    html += """    </tbody>
  </table>
</div>
</body></html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"   HTML guardado: {output_path}")


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--correr", action="store_true", help="Correr el modelo para cada combinacion")
    args = parser.parse_args()

    # Load input
    input_path = project_root / "comparaciones" / "input_comparaciones.yaml"
    with open(input_path, "r", encoding="utf-8") as f:
        input_data = yaml.safe_load(f)

    comparaciones = input_data.get("comparaciones", [])
    print(f"\n{'='*70}")
    print(f"COMPARACION MODELO vs ANALISTA — {len(comparaciones)} corridas")
    print(f"{'='*70}")

    results = []

    for i, comp in enumerate(comparaciones):
        site = comp["site"]
        q_ant = comp["quarter_anterior"]
        q_act = comp["quarter_actual"]
        update_tipo = comp["update_tipo"]
        from nps_model.utils.dates import quarter_fecha_final
        fecha_final = quarter_fecha_final(q_act)

        print(f"\n[{i+1}/{len(comparaciones)}] {site} {update_tipo} {q_ant} vs {q_act}")

        if args.correr:
            print(f"   Actualizando config...")
            update_config(site, q_ant, q_act, update_tipo, comp.get("fecha_corte"))
            print(f"   Corriendo modelo...")
            ok, output = run_model()
            if not ok:
                print(f"   ❌ Error corriendo modelo")
                results.append({
                    **comp,
                    "variacion_nps": None,
                    "explicacion_modelo": "ERROR: modelo fallo",
                })
                continue

        # Extract razonamiento
        print(f"   Extrayendo razonamiento...")

        # Need to reload config for this run
        update_config(site, q_ant, q_act, update_tipo)
        r = load_razonamiento(site, fecha_final)

        if r:
            var_nps = r.get("bloque1", {}).get("var_qvsq")
            parrafo = strip_html(r.get("parrafo_resumen", ""))
            nps_actual = r.get("bloque1", {}).get("nps_actual")
            print(f"   ✅ NPS: {nps_actual:.1f} p.p., var: {var_nps:+.1f}pp" if var_nps else "   ✅ Razonamiento cargado")
        else:
            var_nps = None
            parrafo = ""
            print(f"   ⚠️  Sin datos de razonamiento (correr modelo primero)")

        results.append({
            "site": site,
            "quarter_anterior": q_ant,
            "quarter_actual": q_act,
            "update_tipo": update_tipo,
            "variacion_nps": round(var_nps, 1) if var_nps else None,
            "explicacion_analista": comp.get("explicacion_analista", ""),
            "explicacion_modelo": parrafo,
            "convalida": comp.get("convalida"),
            "comment": comp.get("comment"),
        })

    # Generate outputs
    output_dir = project_root / "comparaciones"
    output_dir.mkdir(exist_ok=True)

    print(f"\n{'='*70}")
    print(f"GENERANDO OUTPUTS")
    print(f"{'='*70}")

    generate_excel(results, output_dir / "comparacion_modelo_vs_analista.xlsx")
    generate_html(results, output_dir / "comparacion_visual.html")

    # Summary
    with_modelo = sum(1 for r in results if r.get("explicacion_modelo"))
    with_analista = sum(1 for r in results if r.get("explicacion_analista"))
    print(f"\n   Resumen: {len(results)} corridas")
    print(f"   Con explicacion modelo: {with_modelo}")
    print(f"   Con explicacion analista: {with_analista}")
    print(f"   Pendientes de validar: {sum(1 for r in results if not r.get('convalida'))}")

    # Restore original config if needed
    if args.correr and comparaciones:
        last = comparaciones[-1]
        # Don't restore — leave last config


if __name__ == "__main__":
    main()
